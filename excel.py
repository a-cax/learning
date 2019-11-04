# program wczytujacy plik excel, wprowadzajacy zmiany i zapisujacy plik pod nowa nazwa

import openpyxl

# sciezka do pliku bazowego
path = "D:\\programowanie\\repozytorium\\python\\learning\\data.xlsx"

# tworzenie zeszytu 
wb = openpyxl.load_workbook(path)
  
# pobranie aktywnego arkusza z zeszytu excel 
sheet = wb.active 
  
# okreslenie adresu komorki do wprowadzenia danych
c1 = sheet.cell(row = 1, column = 4) 
  
# przypisanie wartosci do wybranej wczesniej komorki
c1.value = "maz"
  
c2 = sheet.cell(row = 2 , column = 4) 
c2.value = "zona"
  
# inny sposob wywowalania adresu komorki dla aktywnego arkusza 
c3 = sheet['E1'] 
c3.value = "Anny"
  
c4 = sheet['E2'] 
c4.value = "Piotra"
  
#zapisanie arkusza po wprowadzeniu zmian (w nawiasie podana sciezka i nazwa nowego pliku)
wb.save("D:\\programowanie\\repozytorium\\python\\learning\\data_test.xlsx")
